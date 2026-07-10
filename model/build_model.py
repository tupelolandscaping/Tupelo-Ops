#!/usr/bin/env python3
"""Generate the financial model workbook from the five ledger CSVs.

Usage:
    python model/build_model.py

Reads model/data/ledger-{revenue,labor,overhead,materials,capital}.csv
(identical 11-column schema) and model/data/assumptions.csv, and generates
model/financial-model.xlsx (gitignored, disposable build output, per D7)
with one sheet per derived view, per CONTEXT.md Section 9:

  - Monthly P&L      -- Billed and Collected shown side by side, never
                        merged (D5a); every category decomposed into its
                        ACTUAL and ESTIMATE amounts, with BLOCKED categories
                        noted explicitly rather than shown as a bare $0.
  - Plan vs Actual    -- built (H-052) per model/PROJECTION-PLAN.md and
                        model/data/assumptions.csv (a dated, sourced,
                        labeled-assumption-cell file -- D6-compliant, never a
                        hardcoded number in this script). Shows actual months
                        computed identically to the Monthly P&L sheet, then
                        projects forward in three scenarios (low: flat-
                        continue the last full month's net; mid: full
                        seasonal-cycle average net; high: seasonal-naive
                        revenue growth with Konji's phase-switch and the
                        Xavier/owner-truck-debt triggers applied as discrete
                        events against the running projected balance) rather
                        than a single point estimate -- the underlying
                        history (13 months of billed data, one full seasonal
                        cycle) can't responsibly support more precision than
                        that. The peak-revenue figure is derived fresh from
                        the ledger every build, never stored, specifically so
                        it cannot go stale the way the $4,908 citation in
                        CONTEXT.md did (corrected same pass, H-052).
  - Quarterly Rollups -- calendar quarters (Jan-Mar/Apr-Jun/Jul-Sep/Oct-Dec),
                        summed directly from the monthly P&L figures.
  - Revenue by Service / Revenue by Customer -- built only from
                        event=invoice rows; payment rows carry no service
                        classification (their subcategory is a match-
                        confidence tag, not a service name).
  - Reconciliation    -- scoped to what the ledger actually claims to cover:
                        revenue invoice total vs. the known $28,316.57
                        anchor; revenue payment classification completeness
                        (per H-044); overhead ACTUAL rows vs. their sourced
                        Relay matches; capital rows vs. their sourced
                        transactions; labor ACTUAL rows vs. Relay payroll
                        NET/TAX cash-outs, by reference to
                        model/reconcile_payroll_relay.py's own full-history
                        gate (per H-049), not re-run inline here. States on
                        its face what it does NOT verify (materials' BLOCKED
                        placeholder, the ~85% of real Spend-side Relay cash
                        that has no ledger row at all, and labor's own known
                        residual gaps -- Follow-Ups #19-20) rather than
                        implying full coverage.
  - Raw Ledger        -- all five files unioned as one passthrough table,
                        sorted by date, `type` column distinguishing rows.

Blended labor rate is computed at build time (sum(amount)/sum(quantity))
wherever shown -- never a stored constant. Scoped to crew roles (Crew Lead,
Crew Member) only, per CREW_ROLES -- since H-049 sourced labor from real
payroll data, ledger-labor.csv also carries CEO wage draws and a Landscape
Consultant role, neither a crew staffing cost; blending them in would
distort the figure this KPI exists to inform (future crew-job cost
estimates). Non-crew labor dollars are still shown, in their own column,
not silently dropped.

Re-running this script is idempotent: identical input CSVs produce a
byte-identical workbook (no timestamps or non-deterministic ordering).
"""
import csv
import re
import zipfile
from collections import defaultdict
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font

# openpyxl's writer unconditionally re-stamps docProps/core.xml's `modified`
# field to datetime.now() inside save() (openpyxl/writer/excel.py), which
# can't be overridden by setting wb.properties beforehand -- it would make
# the workbook non-idempotent even when the CSV inputs are unchanged. Fixed
# by post-processing the saved .xlsx: rewrite docProps/core.xml with a fixed
# timestamp and rebuild the zip with fixed per-entry metadata, so byte-for-
# byte output depends only on the CSV inputs, never on wall-clock save time.
FIXED_TIMESTAMP = "2000-01-01T00:00:00Z"
FIXED_ZIP_DATE_TIME = (1980, 1, 1, 0, 0, 0)

LEDGER_FILES = {
    "revenue": "model/data/ledger-revenue.csv",
    "labor": "model/data/ledger-labor.csv",
    "overhead": "model/data/ledger-overhead.csv",
    "materials": "model/data/ledger-materials.csv",
    "capital": "model/data/ledger-capital.csv",
}
OUTPUT_PATH = "model/financial-model.xlsx"

REVENUE_INVOICE_ANCHOR = 28316.57  # $27,891.65 gross + $424.92 tips, H-043

# Crew roles for the blended-labor-rate KPI, per model/data/employee-role-map.csv
# (H-049). Deliberately excludes CEO and Landscape Consultant -- see
# build_monthly_pnl's blended_rate comment for why.
CREW_ROLES = {"Crew Lead", "Crew Member"}


def load_ledgers():
    ledgers = {}
    for name, path in LEDGER_FILES.items():
        with open(path, newline="") as f:
            rows = list(csv.DictReader(f))
        for r in rows:
            r["amount"] = float(r["amount"])
            r["quantity"] = float(r["quantity"]) if r["quantity"] else 0.0
        ledgers[name] = rows
    return ledgers


def year_month(d):
    return d[:7]


def all_months(ledgers):
    months = set()
    for rows in ledgers.values():
        for r in rows:
            months.add(year_month(r["date"]))
    return sorted(months)


def split_actual_estimate(rows):
    actual = sum(r["amount"] for r in rows if r["status"] == "ACTUAL")
    estimate = sum(r["amount"] for r in rows if r["status"] == "ESTIMATE")
    blocked = [r for r in rows if r["status"] == "BLOCKED"]
    return actual, estimate, blocked


# ---------------------------------------------------------------------------
# Monthly P&L
# ---------------------------------------------------------------------------

def build_monthly_pnl(ws, ledgers, months):
    header = [
        "Month",
        "Billed (ACTUAL)", "Billed (ESTIMATE)",
        "Collected (ACTUAL)", "Collected (ESTIMATE)",
        "Labor (ACTUAL)", "Labor (ESTIMATE)",
        "Overhead (ACTUAL)", "Overhead (ESTIMATE)",
        "Materials", "Materials note",
        "Capital (ACTUAL)", "Capital (ESTIMATE)",
        "Blended crew labor rate ($/hr)",
        "Non-crew labor $ (CEO/Consultant, excluded from blended rate)",
        "Note",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for m in months:
        rev = [r for r in ledgers["revenue"] if year_month(r["date"]) == m]
        inv = [r for r in rev if r["event"] == "invoice"]
        pay = [r for r in rev if r["event"] == "payment"]
        labor = [r for r in ledgers["labor"] if year_month(r["date"]) == m]
        overhead = [r for r in ledgers["overhead"] if year_month(r["date"]) == m]
        materials = [r for r in ledgers["materials"] if year_month(r["date"]) == m]
        capital = [r for r in ledgers["capital"] if year_month(r["date"]) == m]

        inv_a, inv_e, _ = split_actual_estimate(inv)
        pay_a, pay_e, _ = split_actual_estimate(pay)
        labor_a, labor_e, _ = split_actual_estimate(labor)
        oh_a, oh_e, _ = split_actual_estimate(overhead)
        cap_a, cap_e, _ = split_actual_estimate(capital)

        mat_a, mat_e, mat_blocked = split_actual_estimate(materials)
        if mat_blocked:
            mat_display = "BLOCKED"
            mat_note = f"{len(mat_blocked)} placeholder row(s), $0.00 -- no per-job data yet (D1)"
        elif materials:
            mat_display = mat_a + mat_e
            mat_note = ""
        else:
            mat_display = "BLOCKED"
            mat_note = "no data this month -- materials/fuel not yet tracked (D1)"

        # Blended rate is scoped to crew roles (Crew Lead, Crew Member) only, per
        # D2's original intent (a crew staffing-cost KPI) -- since H-049 sourced
        # labor from real payroll data, ledger-labor.csv now also carries CEO wage
        # draws and a Landscape Consultant role, neither a crew staffing cost;
        # blending them in would distort the figure this KPI exists to inform
        # (future crew-job cost estimates), not just add noise to it.
        crew_labor = [r for r in labor if r["subcategory"] in CREW_ROLES]
        noncrew_labor = [r for r in labor if r["subcategory"] not in CREW_ROLES]
        crew_hours = sum(r["quantity"] for r in crew_labor)
        crew_a, crew_e, _ = split_actual_estimate(crew_labor)
        blended_rate = (crew_a + crew_e) / crew_hours if crew_hours else None
        noncrew_a, noncrew_e, _ = split_actual_estimate(noncrew_labor)
        noncrew_amount = noncrew_a + noncrew_e

        note = ""
        if m == "2026-07" and labor_a == 0 and labor_e == 0:
            note = ("Re-verified 2026-07-10 (H-049), not a stale artifact: labor is now sourced "
                    "from real Gusto payroll data, dated by period-start (accrual convention). "
                    "The last completed real pay period is 06/15-06/28/2026 (paid 07/07/2026); "
                    "the next period (06/29-07/12/2026) hadn't been processed/paid yet when the "
                    "payroll export was pulled (export window ends 07/09/2026), so it has no row "
                    "at all -- labor is correctly $0 in July here, because no July-dated pay "
                    "period has completed yet, not because data is missing.")

        ws.append([
            m,
            round(inv_a, 2), round(inv_e, 2),
            round(pay_a, 2), round(pay_e, 2),
            round(labor_a, 2), round(labor_e, 2),
            round(oh_a, 2), round(oh_e, 2),
            mat_display, mat_note,
            round(cap_a, 2), round(cap_e, 2),
            round(blended_rate, 2) if blended_rate is not None else "",
            round(noncrew_amount, 2),
            note,
        ])


# ---------------------------------------------------------------------------
# Plan vs Actual & Cash Trajectory (per model/PROJECTION-PLAN.md, H-052)
# ---------------------------------------------------------------------------

ASSUMPTIONS_PATH = "model/data/assumptions.csv"


def load_assumptions():
    with open(ASSUMPTIONS_PATH, newline="") as f:
        lines = [line for line in f if not line.startswith("#")]
    rows = list(csv.DictReader(lines))
    for r in rows:
        r["value"] = float(r["value"]) if r["value"] else 0.0
    return rows


def get_assumption(assumptions, assumption_id, for_month=None):
    """Return the assumption row for `assumption_id` active in `for_month`
    (YYYY-MM), selecting among time-varying rows by effective_date/end_date --
    same pattern as model/populate_labor_from_payroll.py's role_for(). Blank
    effective_date means "from the start"; blank end_date means "open-ended".
    With for_month=None, expects exactly one row (a non-time-varying
    assumption) and returns it directly."""
    candidates = [r for r in assumptions if r["assumption_id"] == assumption_id]
    if for_month is None:
        if len(candidates) != 1:
            raise ValueError(f"{assumption_id}: expected exactly one row for a "
                              f"non-time-varying lookup, got {len(candidates)}")
        return candidates[0]
    matches = [r for r in candidates
               if (not r["effective_date"] or for_month >= r["effective_date"][:7])
               and (not r["end_date"] or for_month <= r["end_date"][:7])]
    if not matches:
        return None
    matches.sort(key=lambda r: r["effective_date"] or "0000-00")
    return matches[-1]


def month_add(ym, n):
    y, m = map(int, ym.split("-"))
    m += n
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return f"{y:04d}-{m:02d}"


def build_plan_vs_actual(ws, ledgers, months, assumptions):
    ws.append(["Plan vs Actual & Cash Trajectory"])
    ws["A1"].font = Font(bold=True)
    ws.append([])

    def collected_amount(m):
        rows = [r for r in ledgers["revenue"] if year_month(r["date"]) == m and r["event"] == "payment"]
        a, e, _ = split_actual_estimate(rows)
        return a + e

    def billed_amount(m):
        rows = [r for r in ledgers["revenue"] if year_month(r["date"]) == m and r["event"] == "invoice"]
        a, e, _ = split_actual_estimate(rows)
        return a + e

    def total_labor_amount(m):
        rows = [r for r in ledgers["labor"] if year_month(r["date"]) == m]
        a, e, _ = split_actual_estimate(rows)
        return a + e

    def crew_labor_amount(m):
        rows = [r for r in ledgers["labor"] if year_month(r["date"]) == m and r["subcategory"] in CREW_ROLES]
        a, e, _ = split_actual_estimate(rows)
        return a + e

    def overhead_amount(m):
        rows = [r for r in ledgers["overhead"] if year_month(r["date"]) == m]
        a, e, _ = split_actual_estimate(rows)
        return a + e

    # --- Peak revenue: derived fresh from the ledger every build, never
    # stored -- this is the exact mechanism that would have caught June 2026's
    # $6,019.89 automatically instead of the $4,908 citation going stale
    # (CONTEXT.md Section 3, corrected H-052).
    billed_by_month = {m: billed_amount(m) for m in months if billed_amount(m) > 0}
    peak_month = max(billed_by_month, key=billed_by_month.get)
    peak_amount = billed_by_month[peak_month]
    peak_target = get_assumption(assumptions, "peak_revenue_target")
    ws.append(["Peak monthly revenue (derived fresh from ledger-revenue.csv, never stored)"])
    ws.append(["  Peak month to date:", peak_month, "  Amount:", round(peak_amount, 2)])
    ws.append(["  Target (assumptions.csv):", peak_target["value"],
                "  Progress:", f"{100 * peak_amount / peak_target['value']:.1f}%"])
    ws.append([])

    # --- Actual months: identical computation to Monthly P&L (Step 5 cross-check) ---
    ws.append(["Actual months -- same computation as the Monthly P&L sheet (collected minus labor minus overhead)"])
    ws.append(["Month", "Collected", "Labor", "Overhead", "Net"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for m in months:
        collected = collected_amount(m)
        labor_total = total_labor_amount(m)
        oh_total = overhead_amount(m)
        net = collected - labor_total - oh_total
        ws.append([m, round(collected, 2), round(labor_total, 2), round(oh_total, 2), round(net, 2)])
    last_real_month = months[-1]
    ws.append([])

    # --- Assumptions driving the projection ---
    growth_rate_row = get_assumption(assumptions, "revenue_growth_rate_assumption")
    growth_rate = growth_rate_row["value"] / 100.0
    low_base_month = month_add(last_real_month, -1)  # last full calendar month
    trailing_3 = [month_add(low_base_month, -2), month_add(low_base_month, -1), low_base_month]
    overhead_trend = sum(overhead_amount(m) for m in trailing_3) / 3

    # Crew labor: seasonal-naive (prior-year same-month x growth rate), NOT a
    # flat trailing-3-month trend. Investigated and confirmed, not assumed:
    # crew labor correlates with revenue seasonally (r=0.77 vs billed revenue,
    # r=0.69 vs collected, over the cleanest 13-month window with full data
    # coverage on both sides) -- the starkest evidence is that crew labor and
    # revenue both go to exactly $0 in Jan/Feb 2026, confirmed NOT a hiring-gap
    # artifact (Konji's first labor row is 2025-05-19, well before this gap).
    # A flat trend would wrongly carry ~$1,870/mo of crew cost into a
    # projected January or February, when the real historical pattern shows
    # it collapsing to zero. Overhead stays flat-continued (fixed monthly
    # contracts, not seasonal -- Jan/Feb 2026 overhead was $281.75, nonzero,
    # confirming it does NOT follow the same pattern).
    crew_growth_rate_row = get_assumption(assumptions, "crew_labor_growth_rate_assumption")
    crew_growth_rate = crew_growth_rate_row["value"] / 100.0

    reserve = get_assumption(assumptions, "cash_buffer_reserve")["value"]
    truck_hedge = get_assumption(assumptions, "cash_buffer_truck_hedge")["value"]
    buffer_target = reserve + truck_hedge
    cash_start = get_assumption(assumptions, "cash_starting_balance")
    xavier = get_assumption(assumptions, "xavier_payout")
    owner_debt = get_assumption(assumptions, "owner_truck_debt_repayment")

    ws.append(["Projection assumptions (model/data/assumptions.csv; recomputed fresh from the ledger each build, per D6 -- never a stored constant)"])
    ws.append(["  Revenue growth-rate assumption:", f"{growth_rate_row['value']}%",
                "  source:", growth_rate_row["source"]])
    ws.append(["  Crew labor: seasonal-naive (prior-year same-month x growth rate), NOT a flat trend --"])
    ws.append(["    see caveats below. Crew-labor growth-rate assumption:", f"{crew_growth_rate_row['value']}%",
                "  OPEN DECISION, see crew_growth_rate_row's own label"])
    ws.append(["  Overhead trend (flat, trailing 3 real months, " + ", ".join(trailing_3) + "):",
                round(overhead_trend, 2)])
    ws.append(["  Cash-buffer target:", buffer_target, "  Starting cash:", cash_start["value"],
                f"(as of {cash_start['effective_date']}, {cash_start['source']})"])
    ws.append([])
    ws.append(["Caveats -- read before trusting any number below:"])
    ws.append(["  - Crew labor is projected seasonal-naive (prior-year same month x a growth-rate"])
    ws.append(["    assumption), confirmed correlated with revenue (r=0.77 vs billed, cleanest 13-month"])
    ws.append(["    window) -- NOT a flat trend. Which growth rate to use is an OPEN DECISION: crew"])
    ws.append(["    labor's own observed YoY growth (92.8%) differs materially from revenue's (75.9%,"])
    ws.append(["    reflecting Konji's staffing ramp-up as a driver beyond volume growth); this build"])
    ws.append(["    uses crew labor's own rate as the default, not yet owner-confirmed."])
    ws.append(["  - Anais's canvassing/backup-lead cost (BLOCKED, Follow-Up #11) is excluded entirely,"])
    ws.append(["    not summed as $0 -- her ordinary Crew Member wages ARE already in the crew labor line."])
    ws.append(["  - Konji's canvassing bonuses and winter-project rate (assumptions.csv) are NOT included"])
    ws.append(["    below -- no client-volume/hours assumption exists yet to multiply them against."])
    ws.append(["  - Konji's 6% revenue share is applied to PROJECTED COLLECTED cash as a simplification;"])
    ws.append(["    the real terms (D3/H-031/H-032) compute it on NET BILLED revenue for jobs he leads"])
    ws.append(["    specifically, which this projection does not separately track."])
    ws.append(["  - cash_starting_balance is a dated snapshot, not re-derivable from the ledger (~85% of"])
    ws.append(["    real Spend-side Relay cash has no ledger row) -- needs periodic re-verification,"])
    ws.append(["    per CONTEXT.md Follow-Up #21."])
    ws.append([])

    # --- Projection horizon: month after the last real month, through the
    # plan's stated horizon (June 2027, matching D4's phase-switch boundary).
    horizon_end = "2027-06"
    projected_months = []
    cur = month_add(last_real_month, 1)
    while cur <= horizon_end:
        projected_months.append(cur)
        cur = month_add(cur, 1)

    gap = buffer_target - cash_start["value"]

    # --- Low scenario: flat-continue the last full calendar month's actual net. ---
    low_net = collected_amount(low_base_month) - total_labor_amount(low_base_month) - overhead_amount(low_base_month)
    low_months_to_buffer = "Never at this rate (negative)" if low_net <= 0 else round(gap / low_net, 1)

    # --- Mid scenario: full seasonal-cycle average net (trailing 12 real months). ---
    cycle_months = [month_add(low_base_month, -i) for i in range(11, -1, -1)]
    cycle_net = sum(collected_amount(m) - total_labor_amount(m) - overhead_amount(m) for m in cycle_months)
    mid_net = cycle_net / 12
    mid_months_to_buffer = "Never at this rate (negative)" if mid_net <= 0 else round(gap / mid_net, 1)

    # --- High scenario: seasonal-naive growth-adjusted, month-by-month, with
    # Konji's phase-switch and Xavier/owner triggers applied as discrete
    # events -- computed once here so both the summary line and the detailed
    # table below are derived from the same run, not duplicated logic. ---
    high_rows = []
    balance = cash_start["value"]
    xavier_paid = False
    owner_paid = False
    buffer_reached_month = None
    for m in projected_months:
        prior_year_month = month_add(m, -12)
        proj_revenue = collected_amount(prior_year_month) * (1 + growth_rate)
        proj_crew_labor = crew_labor_amount(prior_year_month) * (1 + crew_growth_rate)

        share_row = get_assumption(assumptions, "konji_revenue_share_pct", for_month=m)
        konji_share = proj_revenue * (share_row["value"] / 100.0) if share_row else 0.0

        proj_outflow = proj_crew_labor + overhead_trend + konji_share
        net = proj_revenue - proj_outflow
        balance_start = balance
        balance_before_payout = balance_start + net

        payout_label = ""
        payout_amount = 0.0
        if not xavier_paid and balance_before_payout >= buffer_target:
            payout_label = "Xavier payout"
            payout_amount = xavier["value"]
            xavier_paid = True
            buffer_reached_month = m
        elif xavier_paid and not owner_paid and balance_before_payout >= buffer_target:
            payout_label = "Owner truck-debt repayment"
            payout_amount = owner_debt["value"]
            owner_paid = True

        balance = balance_before_payout - payout_amount
        high_rows.append((m, proj_revenue, proj_crew_labor, konji_share, proj_outflow, net,
                           balance_start, payout_label, payout_amount, balance))

    if buffer_reached_month:
        months_elapsed = projected_months.index(buffer_reached_month) + 1
        high_summary = f"Reaches $6,000 in {buffer_reached_month} (~{months_elapsed} months)"
    else:
        high_summary = f"Not reached within the horizon (ends {projected_months[-1]} at ${balance:,.2f})"

    ws.append(["Scenario summary -- a range, not a single point projection (thin history: "
               "13 months of billed data, one full seasonal cycle)"])
    ws.append(["Scenario", "Basis", "Monthly net rate", "Months to $6,000 buffer"])
    ws.append(["Low", f"Flat-continue {low_base_month}'s actual net", round(low_net, 2), low_months_to_buffer])
    ws.append(["Mid", f"Full seasonal-cycle average ({cycle_months[0]}..{cycle_months[-1]})",
               round(mid_net, 2), mid_months_to_buffer])
    ws.append(["High", "Growth-adjusted, month-by-month, seasonal-naive crew labor (see table below)",
               "", high_summary])
    ws.append([])

    ws.append(["High scenario -- month-by-month (seasonal-naive revenue, seasonal-naive crew labor, "
               "flat overhead trend, Konji's phase-switch and Xavier/owner triggers applied as discrete events)"])
    ws.append(["Month", "Projected revenue", "Projected crew labor (seasonal-naive)", "Konji share",
                "Projected outflow", "Net", "Balance (start)", "Payout this month", "Balance (end)"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for (m, proj_revenue, proj_crew_labor, konji_share, proj_outflow, net,
         balance_start, payout_label, payout_amount, balance_end) in high_rows:
        ws.append([m, round(proj_revenue, 2), round(proj_crew_labor, 2), round(konji_share, 2),
                   round(proj_outflow, 2), round(net, 2), round(balance_start, 2),
                   f"{payout_label} (${payout_amount:,.2f})" if payout_label else "",
                   round(balance_end, 2)])


# ---------------------------------------------------------------------------
# Quarterly Rollups
# ---------------------------------------------------------------------------

def quarter_of(ym):
    y, m = ym.split("-")
    m = int(m)
    q = (m - 1) // 3 + 1
    return f"{y}-Q{q}"


def build_quarterly_rollups(ws, ledgers, months):
    header = [
        "Quarter",
        "Billed (ACTUAL)", "Billed (ESTIMATE)",
        "Collected (ACTUAL)", "Collected (ESTIMATE)",
        "Labor (ACTUAL)", "Labor (ESTIMATE)",
        "Overhead (ACTUAL)", "Overhead (ESTIMATE)",
        "Capital (ACTUAL)", "Capital (ESTIMATE)",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    quarters = defaultdict(lambda: defaultdict(float))
    for m in months:
        q = quarter_of(m)
        rev = [r for r in ledgers["revenue"] if year_month(r["date"]) == m]
        inv = [r for r in rev if r["event"] == "invoice"]
        pay = [r for r in rev if r["event"] == "payment"]
        labor = [r for r in ledgers["labor"] if year_month(r["date"]) == m]
        overhead = [r for r in ledgers["overhead"] if year_month(r["date"]) == m]
        capital = [r for r in ledgers["capital"] if year_month(r["date"]) == m]

        inv_a, inv_e, _ = split_actual_estimate(inv)
        pay_a, pay_e, _ = split_actual_estimate(pay)
        labor_a, labor_e, _ = split_actual_estimate(labor)
        oh_a, oh_e, _ = split_actual_estimate(overhead)
        cap_a, cap_e, _ = split_actual_estimate(capital)

        quarters[q]["inv_a"] += inv_a
        quarters[q]["inv_e"] += inv_e
        quarters[q]["pay_a"] += pay_a
        quarters[q]["pay_e"] += pay_e
        quarters[q]["labor_a"] += labor_a
        quarters[q]["labor_e"] += labor_e
        quarters[q]["oh_a"] += oh_a
        quarters[q]["oh_e"] += oh_e
        quarters[q]["cap_a"] += cap_a
        quarters[q]["cap_e"] += cap_e

    for q in sorted(quarters):
        v = quarters[q]
        ws.append([
            q,
            round(v["inv_a"], 2), round(v["inv_e"], 2),
            round(v["pay_a"], 2), round(v["pay_e"], 2),
            round(v["labor_a"], 2), round(v["labor_e"], 2),
            round(v["oh_a"], 2), round(v["oh_e"], 2),
            round(v["cap_a"], 2), round(v["cap_e"], 2),
        ])


# ---------------------------------------------------------------------------
# Revenue by Service / Revenue by Customer (invoice rows only)
# ---------------------------------------------------------------------------

def build_revenue_by_service(ws, ledgers):
    ws.append(["Category", "Subcategory (service)", "Amount (ACTUAL)", "Amount (ESTIMATE)", "Row count"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    inv = [r for r in ledgers["revenue"] if r["event"] == "invoice"]
    grouped = defaultdict(lambda: {"actual": 0.0, "estimate": 0.0, "count": 0})
    for r in inv:
        key = (r["category"], r["subcategory"])
        if r["status"] == "ACTUAL":
            grouped[key]["actual"] += r["amount"]
        elif r["status"] == "ESTIMATE":
            grouped[key]["estimate"] += r["amount"]
        grouped[key]["count"] += 1

    for (cat, subcat) in sorted(grouped, key=lambda k: -grouped[k]["actual"]):
        v = grouped[(cat, subcat)]
        ws.append([cat, subcat, round(v["actual"], 2), round(v["estimate"], 2), v["count"]])


def build_revenue_by_customer(ws, ledgers):
    ws.append(["Customer", "Amount (ACTUAL)", "Amount (ESTIMATE)", "Row count"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    inv = [r for r in ledgers["revenue"] if r["event"] == "invoice" and r["customer"]]
    grouped = defaultdict(lambda: {"actual": 0.0, "estimate": 0.0, "count": 0})
    for r in inv:
        key = r["customer"]
        if r["status"] == "ACTUAL":
            grouped[key]["actual"] += r["amount"]
        elif r["status"] == "ESTIMATE":
            grouped[key]["estimate"] += r["amount"]
        grouped[key]["count"] += 1

    for cust in sorted(grouped, key=lambda k: -grouped[k]["actual"]):
        v = grouped[cust]
        ws.append([cust, round(v["actual"], 2), round(v["estimate"], 2), v["count"]])


# ---------------------------------------------------------------------------
# Reconciliation -- scoped to what's actually checkable
# ---------------------------------------------------------------------------

def build_reconciliation(ws, ledgers):
    ws.append(["Reconciliation"])
    ws["A1"].font = Font(bold=True)
    ws.append([])

    inv = [r for r in ledgers["revenue"] if r["event"] == "invoice"]
    inv_total = sum(r["amount"] for r in inv)
    check1_pass = abs(inv_total - REVENUE_INVOICE_ANCHOR) < 0.01
    ws.append(["Check 1: revenue invoice total vs. known anchor"])
    ws.append(["  Ledger total:", round(inv_total, 2)])
    ws.append(["  Anchor ($27,891.65 gross + $424.92 tips, H-043):", REVENUE_INVOICE_ANCHOR])
    ws.append(["  Result:", "PASS" if check1_pass else "FAIL"])
    ws.append([])
    if not check1_pass:
        print(f"GATE FAILED: Check 1 -- revenue invoice total ${inv_total:,.2f} does not match "
              f"the ${REVENUE_INVOICE_ANCHOR:,.2f} anchor (H-043), off by "
              f"${inv_total - REVENUE_INVOICE_ANCHOR:,.2f}. Not writing the workbook.")
        raise SystemExit(1)

    pay = [r for r in ledgers["revenue"] if r["event"] == "payment"]
    pay_total = sum(r["amount"] for r in pay)
    ws.append(["Check 2: revenue payment classification completeness (per H-044)"])
    ws.append(["  Payment events in ledger:", len(pay), "  sum:", round(pay_total, 2)])
    ws.append(["  H-044's classification gate already proved: 94 settled Receive/"])
    ws.append(["  Receive-transfer rows (deduplicated) = 58 classified payment events"])
    ws.append(["  + 36 named exclusions, zero residual. Re-verifying that gate is not"])
    ws.append(["  repeated here -- see model/match_payments.py's own gate output."])
    ws.append(["  Result:", "PASS (by reference to H-044's gate)"])
    ws.append([])

    overhead = ledgers["overhead"]
    oh_actual = [r for r in overhead if r["status"] == "ACTUAL"]
    ws.append(["Check 3: overhead ACTUAL rows sourced to real Relay transactions"])
    ws.append(["  ACTUAL overhead rows:", len(oh_actual), "  sum:", round(sum(r["amount"] for r in oh_actual), 2)])
    ws.append(["  Every row's `source` field cites a specific reference/Relay*.csv file and"])
    ws.append(["  transaction reference (H-046) -- spot-checkable row by row, not re-verified"])
    ws.append(["  in aggregate here."])
    ws.append(["  Result:", "PASS (sourced; not independently re-matched here)"])
    ws.append([])

    capital = ledgers["capital"]
    ws.append(["Check 4: capital rows sourced to real Relay transactions"])
    ws.append(["  Capital rows:", len(capital), "  sum:", round(sum(r["amount"] for r in capital), 2)])
    for r in capital:
        ws.append(["   -", r["date"], r["category"], r["subcategory"], r["amount"], r["source"][:80]])
    ws.append(["  Result:", "PASS (both rows trace to a specific, named transaction)"])
    ws.append([])

    labor = ledgers["labor"]
    labor_actual = [r for r in labor if r["status"] == "ACTUAL"]
    ws.append(["Check 5: labor ACTUAL rows reconciled against Relay payroll cash-outs (per H-049)"])
    ws.append(["  ACTUAL labor rows:", len(labor_actual), "  sum:", round(sum(r["amount"] for r in labor_actual), 2)])
    ws.append(["  model/reconcile_payroll_relay.py's own full-history gate already proved: of"])
    ws.append(["  25 non-reversed, non-empty pay periods, the 24 with nonzero Net Pay matched a"])
    ws.append(["  specific Relay GUSTO NET transaction exactly (24/24), and 24 of the 25 with"])
    ws.append(["  nonzero combined tax matched a specific Relay GUSTO TAX transaction exactly"])
    ws.append(["  (24/25 -- the one exception is a $5.37 employer-tax-only correction with no"])
    ws.append(["  discrete Relay transaction, Follow-Up #20). Re-verifying that gate is not"])
    ws.append(["  repeated here -- see model/reconcile_payroll_relay.py's own gate output."])
    ws.append(["  Result:", "PASS (by reference to reconcile_payroll_relay.py's gate; 1 known exception, Follow-Up #20)"])
    ws.append([])

    ws.append(["What this tab does NOT verify:"])
    ws.append(["  - Labor's own known residual gaps (see Check 5 above for what IS verified):"])
    ws.append(["    payroll actuals carry no per-job/per-customer attribution and cover all"])
    ws.append(["    work combined, not recurring-only (Follow-Up #19); the $5.37 correction"])
    ws.append(["    with no discrete Relay match (Follow-Up #20)."])
    ws.append(["  - Materials: BLOCKED placeholder only ($0.00). No per-job cost data exists."])
    ws.append(["  - Commercial Auto and equipment-maintenance overhead lines: BLOCKED, no row"])
    ws.append(["    at all (no confirmed cash date / no allocation basis)."])
    ws.append(["  - The full per-account balance-delta gate (sum of ALL ledger events ="])
    ws.append(["    Relay ending balance) is NOT attempted here: ~85% of real settled"])
    ws.append(["    Spend-side Relay cash ($44,164.11 of $51,738.61) has no ledger row at"])
    ws.append(["    all by design (real Gusto payroll NET/TAX, internal transfers, and"])
    ws.append(["    uncaptured small vendor spend) -- attempting that gate would fail"])
    ws.append(["    predictably, not diagnostically, so it is not presented as a check here."])


# ---------------------------------------------------------------------------
# Raw Ledger passthrough
# ---------------------------------------------------------------------------

def build_raw_ledger(ws, ledgers):
    header = ["date", "type", "event", "category", "subcategory", "customer",
              "quantity", "unit_rate", "amount", "status", "source"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    all_rows = []
    for rows in ledgers.values():
        all_rows.extend(rows)
    all_rows.sort(key=lambda r: (r["date"], r["type"]))

    for r in all_rows:
        ws.append([r["date"], r["type"], r["event"], r["category"], r["subcategory"],
                   r["customer"], r["quantity"], r["unit_rate"], r["amount"],
                   r["status"], r["source"]])


CORE_XML_DATE_RE = re.compile(
    r'(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)'
)


def normalize_xlsx_for_determinism(path):
    """Rewrite the saved workbook so re-running with unchanged input produces
    a byte-identical file: fix docProps/core.xml's timestamps (openpyxl's
    writer stamps `modified` with datetime.now() unconditionally on save) and
    give every zip entry a fixed date_time, in a fixed entry order.
    """
    with zipfile.ZipFile(path, "r") as zin:
        entries = []
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "docProps/core.xml":
                text = data.decode("utf-8")
                text = CORE_XML_DATE_RE.sub(rf"\g<1>{FIXED_TIMESTAMP}\g<2>", text)
                data = text.encode("utf-8")
            entries.append((info.filename, data))

    entries.sort(key=lambda e: e[0])
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for filename, data in entries:
            info = zipfile.ZipInfo(filename, date_time=FIXED_ZIP_DATE_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            zout.writestr(info, data)


def main():
    ledgers = load_ledgers()
    months = all_months(ledgers)
    assumptions = load_assumptions()

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "model/build_model.py"

    build_monthly_pnl(wb.create_sheet("Monthly P&L"), ledgers, months)
    build_plan_vs_actual(wb.create_sheet("Plan vs Actual"), ledgers, months, assumptions)
    build_quarterly_rollups(wb.create_sheet("Quarterly Rollups"), ledgers, months)
    build_revenue_by_service(wb.create_sheet("Revenue by Service"), ledgers)
    build_revenue_by_customer(wb.create_sheet("Revenue by Customer"), ledgers)
    build_reconciliation(wb.create_sheet("Reconciliation"), ledgers)
    build_raw_ledger(wb.create_sheet("Raw Ledger"), ledgers)

    wb.save(OUTPUT_PATH)
    normalize_xlsx_for_determinism(OUTPUT_PATH)
    print(f"[build] wrote {OUTPUT_PATH}")
    print(f"[build] months covered: {months[0]} to {months[-1]} ({len(months)} months)")
    for name, rows in ledgers.items():
        print(f"[build] {name}: {len(rows)} rows")


if __name__ == "__main__":
    main()
